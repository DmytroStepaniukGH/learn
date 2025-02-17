import base64
import openpyxl
from io import BytesIO

from odoo import models, fields, _
from odoo.exceptions import UserError


class ContactImportWizard(models.TransientModel):
    _name = 'contact.import.wizard'
    _description = 'Import Contacts from Excel'

    file_data = fields.Binary("Excel File", required=True)
    file_name = fields.Char("File Name")

    def action_import_contacts(self):
        if not self.file_data:
            raise UserError(_("Please upload an Excel file."))

        data = base64.b64decode(self.file_data)
        workbook = openpyxl.load_workbook(BytesIO(data), data_only=True)

        sheet = workbook.active
        header = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]

        try:
            name_idx = header.index("Назва контакту")
            phone_idx = header.index("Телефон")
            email_idx = header.index("Пошта")

        except ValueError:
            raise UserError(_("The uploaded file does not contain the required columns: "
                              "'Назва контакту', 'Телефон', 'Пошта'."))

        partner_obj = self.env['res.partner']

        for row in sheet.iter_rows(min_row=2, values_only=True):
            name = row[name_idx]
            phone = row[phone_idx]
            email = row[email_idx]
            if name:
                partner_obj.create({
                    'name': name,
                    'phone': phone,
                    'email': email,
                })

        return {'type': 'ir.actions.act_window_close'}